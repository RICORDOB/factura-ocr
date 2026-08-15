"""Capa de persistencia en Excel usando openpyxl.

Reglas de negocio:
- Si el archivo .xlsx NO existe, se crea con las cabeceras definidas.
- Si el archivo YA existe, se abre (load_workbook) y la nueva fila se agrega
  en la primera fila vacía. NUNCA se sobrescribe ni se recrea un archivo existente.

Estructura del libro:
- Hoja "Facturas": una fila por factura (datos de cabecera).
- Hoja "Detalles": una fila por producto, vinculada a su factura por N° Factura.
"""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# Cabeceras en el orden exacto en que se guardan (mismo orden que el frontend)
HEADERS = [
    "Fecha",
    "NIT",
    "Razón Social",
    "N° Factura",
    "Subtotal",
    "IVA",
    "Total",
    "Moneda",
]

# Cabeceras de la hoja de detalle de productos
DETAIL_HEADERS = [
    "N° Factura",
    "Producto",
    "Cantidad",
    "Precio Unitario",
    "Subtotal",
    "IVA",
    "Total",
]

SHEET_FACTURAS = "Facturas"
SHEET_DETALLES = "Detalles"

# Fila donde se escriben las cabeceras (fila 1) y desde dónde parten los datos
HEADER_ROW = 1
FIRST_DATA_ROW = 2

_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_FONT = Font(bold=True, color="FFFFFF")


def _normalize_row(data: dict) -> list:
    """Convierte el dict normalizado del extractor en una fila alineada a HEADERS."""
    return [
        data.get("fecha"),
        data.get("nit"),
        data.get("razon_social"),
        data.get("numero_factura"),
        data.get("subtotal"),
        data.get("iva"),
        data.get("total"),
        data.get("moneda"),
    ]


def _normalize_detail_row(numero_factura: str, item: dict) -> list:
    """Convierte un line_item normalizado en una fila alineada a DETAIL_HEADERS."""
    return [
        numero_factura,
        item.get("producto"),
        item.get("cantidad"),
        item.get("precio_unitario"),
        item.get("subtotal"),
        item.get("iva"),
        item.get("total"),
    ]


def _find_next_empty_row(ws) -> int:
    """Busca la primera fila completamente vacía a partir de FIRST_DATA_ROW."""
    row = FIRST_DATA_ROW
    while True:
        cell = ws.cell(row=row, column=1)
        if cell.value is None or str(cell.value).strip() == "":
            return row
        row += 1


def _style_headers(ws) -> None:
    """Aplica formato simple a la fila de cabeceras (negrita, fondo)."""
    for cell in ws[HEADER_ROW]:
        cell.fill = _FILL
        cell.font = _FONT


def _ensure_sheet(wb, name: str, headers: list[str]):
    """Devuelve la hoja (creándola con cabeceras si no existe)."""
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(title=name)
    ws.append(headers)
    _style_headers(ws)
    return ws


def append_invoice(excel_path: Path, data: dict) -> dict:
    """Agrega una factura (cabecera + sus productos) al archivo Excel.

    Crea el archivo si no existe. Devuelve un dict con el resultado.
    """
    excel_path = Path(excel_path)
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if not excel_path.exists():
        wb = Workbook()
        wb.remove(wb.active)  # elimina la hoja por defecto, se crean las dos
        wb = _create_book(wb)
    else:
        try:
            wb = load_workbook(excel_path)
        except Exception as exc:
            raise RuntimeError(f"No se pudo abrir el archivo Excel existente: {exc}") from exc

    ws_facturas = _ensure_sheet(wb, SHEET_FACTURAS, HEADERS)
    ws_detalles = _ensure_sheet(wb, SHEET_DETALLES, DETAIL_HEADERS)

    target_row = _find_next_empty_row(ws_facturas)
    for col, value in enumerate(_normalize_row(data), start=1):
        ws_facturas.cell(row=target_row, column=col, value=value)

    numero_factura = str(data.get("numero_factura") or "")
    for item in data.get("line_items") or []:
        detail_row = _find_next_empty_row(ws_detalles)
        for col, value in enumerate(_normalize_detail_row(numero_factura, item), start=1):
            ws_detalles.cell(row=detail_row, column=col, value=value)

    try:
        wb.save(excel_path)
    except Exception as exc:
        raise RuntimeError(f"No se pudo guardar el archivo Excel: {exc}") from exc

    return {
        "ok": True,
        "archivo": str(excel_path),
        "fila": target_row,
        "detalle_filas": len(data.get("line_items") or []),
        "fecha_registro": datetime.now().isoformat(timespec="seconds"),
    }


def _create_book(wb: Workbook) -> Workbook:
    """Crea la estructura inicial del libro con ambas hojas y sus cabeceras."""
    ws_facturas = wb.create_sheet(title=SHEET_FACTURAS)
    ws_facturas.append(HEADERS)
    _style_headers(ws_facturas)
    ws_detalles = wb.create_sheet(title=SHEET_DETALLES)
    ws_detalles.append(DETAIL_HEADERS)
    _style_headers(ws_detalles)
    return wb
